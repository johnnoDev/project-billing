import io
from datetime import datetime

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


# ---------------------------------------------------------------------------
# Mixin genérico de exportación (PDF / Excel) para CBV ListView
# ---------------------------------------------------------------------------

class ExportMixin:
    """
    Agrega botones de exportación a cualquier ListView.

    Uso en la vista:
        class MyListView(LoginRequiredMixin, ExportMixin, ListView):
            export_fields = [
                ('name',       'Nombre'),
                ('brand.name', 'Marca'),      # ruta con puntos para FK
                ('is_active',  'Activo'),
                (lambda o: f"${o.unit_price}", 'Precio'),  # callable
            ]
            export_filename = 'productos'
            export_title    = 'Listado de Productos'

    Los botones en el template deben apuntar a la misma URL con
    ?export=pdf  o  ?export=excel  (preservando los demás filtros).
    """

    export_fields: list = []   # [(accessor_or_callable, 'Encabezado'), ...]
    export_filename: str = 'listado'
    export_title: str = 'Listado'

    # -- punto de entrada --------------------------------------------------

    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('export', '').lower()
        if fmt == 'pdf':
            return self._render_pdf(self.get_queryset())
        if fmt == 'excel':
            return self._render_excel(self.get_queryset())
        return super().get(request, *args, **kwargs)

    # -- helpers -----------------------------------------------------------

    @staticmethod
    def _resolve(obj, accessor):
        """Resuelve un accessor (str con puntos o callable) sobre obj."""
        if callable(accessor):
            return accessor(obj)
        value = obj
        for attr in str(accessor).split('.'):
            value = getattr(value, attr, '')
            if callable(value):
                value = value()
        return '' if value is None else str(value)

    def _header_and_rows(self, queryset):
        headers = [label for _, label in self.export_fields]
        rows = [
            [self._resolve(obj, acc) for acc, _ in self.export_fields]
            for obj in queryset
        ]
        return headers, rows

    # -- PDF ---------------------------------------------------------------

    def _render_pdf(self, queryset):
        headers, rows = self._header_and_rows(queryset)
        buf = _build_pdf(self.export_title, headers, rows)
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="{self.export_filename}_{datetime.now():%Y%m%d}.pdf"'
        )
        return response

    # -- Excel -------------------------------------------------------------

    def _render_excel(self, queryset):
        headers, rows = self._header_and_rows(queryset)
        buf = _build_excel(self.export_title, headers, rows)
        response = HttpResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="{self.export_filename}_{datetime.now():%Y%m%d}.xlsx"'
        )
        return response


# ---------------------------------------------------------------------------
# Función auxiliar para vistas basadas en funciones (FBV)
# ---------------------------------------------------------------------------

def export_response(request, queryset, fields, title, filename):
    """
    Genera una respuesta de exportación PDF o Excel para FBVs.

    Parámetros
    ----------
    request  : HttpRequest  (lee GET['export'])
    queryset : QuerySet o lista iterable
    fields   : [(accessor_or_callable, 'Encabezado'), ...]
    title    : str – título del reporte
    filename : str – nombre base del archivo (sin extensión)

    Retorna HttpResponse o None si no hay parámetro de exportación.
    """
    fmt = request.GET.get('export', '').lower()
    if fmt not in ('pdf', 'excel'):
        return None

    def resolve(obj, accessor):
        if callable(accessor):
            return accessor(obj)
        value = obj
        for attr in str(accessor).split('.'):
            value = getattr(value, attr, '')
            if callable(value):
                value = value()
        return '' if value is None else str(value)

    headers = [label for _, label in fields]
    rows = [
        [resolve(obj, acc) for acc, _ in fields]
        for obj in queryset
    ]
    stamp = datetime.now().strftime('%Y%m%d')

    if fmt == 'pdf':
        buf = _build_pdf(title, headers, rows)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}_{stamp}.pdf"'
        return resp

    buf = _build_excel(title, headers, rows)
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}_{stamp}.xlsx"'
    return resp


# ---------------------------------------------------------------------------
# Constructores internos de PDF y Excel
# ---------------------------------------------------------------------------

def _build_pdf(title, headers, rows):
    buf = io.BytesIO()
    num_cols = len(headers)
    page = landscape(A4) if num_cols > 5 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles['Title']),
        Paragraph(
            f"Generado: {datetime.now():%d/%m/%Y %H:%M}  |  {len(rows)} registro(s)",
            styles['Normal'],
        ),
        Spacer(1, 0.4 * cm),
    ]

    table_data = [headers] + rows
    col_width = (doc.width) / num_cols if num_cols else doc.width

    tbl = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),  colors.HexColor('#212529')),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0),  9),
        ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


def _build_excel(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limita a 31 caracteres

    # Fila de título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')

    # Fila de metadatos
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.cell(row=2, column=1,
            value=f"Generado: {datetime.now():%d/%m/%Y %H:%M}  |  {len(rows)} registro(s)")

    ws.append([])  # fila vacía

    # Encabezados
    header_fill = PatternFill(fill_type='solid', fgColor='212529')
    header_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos
    alt_fill = PatternFill(fill_type='solid', fgColor='F8F9FA')
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx % 2 == 0:
                cell.fill = alt_fill

    # Ajuste de ancho de columnas
    for col_cells in ws.iter_cols(min_row=4, max_row=ws.max_row):
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    
    Uso:
        class BrandDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
            ...
    
    ¿POR QUÉ?
    Porque solo el personal autorizado (staff) debe poder
    eliminar registros. Un usuario normal puede ver y crear,
    pero no borrar información importante del sistema.
    
    ¿CÓMO FUNCIONA?
    1. El usuario intenta acceder a una vista protegida
    2. dispatch() se ejecuta ANTES que la vista
    3. Si user.is_staff es False → redirige con mensaje de error
    4. Si user.is_staff es True → ejecuta la vista normalmente
    """

    # URL a donde redirigir si no es staff
    # Se puede sobreescribir en cada vista
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        """
        dispatch() es el primer método que se ejecuta en una CBV.
        Interceptamos aquí para verificar permisos ANTES de
        procesar la petición (GET o POST).
        """
        # Verificar si el usuario es staff
        if not request.user.is_staff:
            # Mostrar mensaje de error al usuario
            messages.error(request, self.staff_error_message)
            # Redirigir a la URL configurada
            return redirect(self.staff_redirect_url)

        # Si es staff, continuar con el flujo normal de la vista
        return super().dispatch(request, *args, **kwargs)
